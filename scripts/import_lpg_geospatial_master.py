"""Build the LPG fallback snapshot from the audited 2026 workbook."""
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import json, re, sys, unicodedata

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"D:\# DOWNLOAD\Master_Geospasial_LPG_Pinrang_Fallback_Lokasi_2026.xlsx")
AGENT_COORD_SOURCE = Path(r"D:\# DOWNLOAD\Tabel_Agen_LPG_3Kg_Kabupaten_Pinrang_Koordinat_Lengkap.xlsx")
VERSION = "2026-08-31-lpg-agent-coordinates-v2"

def records(sheet):
    headers = [cell.value for cell in sheet[1]]
    return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]

def normalized(value):
    raw = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]", "", raw)

def clean_village(value):
    return re.sub(r"^(DESA|KELURAHAN)\s+", "", str(value or "").strip(), flags=re.I).title()

wb = load_workbook(SOURCE, read_only=True, data_only=True)
agent_rows = records(wb["Agen_Juli_2026"])
geo_rows = records(wb["Geospasial_Pangkalan"])

old_agent_ids = {
    "PTGASIFAMULYAPERSADA":"AG-001", "PTHAMISASUKRAHMULYA":"AG-002",
    "PTHABDRAHMANHASYIM":"AG-003", "PTNURCAHAYAENERGIABADI":"AG-004",
    "PTWAHYUDWIKENCANAMANDIRI":"AG-005", "PTNASMANHAFIDMANDIRI":"AG-006",
    "PTHAMIRUDDINRAHMAN":"AG-007", "PTKAKAMIGASUTAMA":"AG-008",
    "PTKIANOENERGIUTAMA":"AG-009",
}

agents = []
for row in agent_rows:
    key = normalized(row["Nama Agen/Penyalur"])
    agents.append({
        "id": old_agent_ids[key], "nationalNumber": int(row["No Nasional"]),
        "name": row["Nama Agen/Penyalur"], "normalizedName": key,
        "address": row["Alamat Resmi"], "status": "ACTIVE",
        "linkedPangkalanCount": int(row["Jumlah Pangkalan Mar 2026"] or 0),
        "relationshipStatus": row["Status Relasi Pangkalan Mar 2026"],
        "sourceType": "DITJEN_MIGAS_OFFICIAL", "sourcePeriod": "Juli 2026",
        "sourceUrl": row["Sumber URL"], "sourceLine": int(row["Ref Line PDF"]),
        "dataVersion": VERSION, "coordinateStatus": "BELUM_DIPETAKAN",
    })

# Koordinat agen merupakan kandidat hasil audit workbook terpisah. Gabungkan
# berdasarkan kode agen; jangan pernah membentuk koordinat agen dari centroid
# pangkalan karena dapat menarik posisi fallback lama.
coord_wb = load_workbook(AGENT_COORD_SOURCE, read_only=True, data_only=True)
coord_ws = coord_wb["Data Agen"]
coord_headers = [cell.value for cell in next(coord_ws.iter_rows(min_row=4, max_row=4))]
coord_rows = [dict(zip(coord_headers, row)) for row in coord_ws.iter_rows(min_row=5, values_only=True) if row[0]]
coord_by_id = {str(row["Kode Agen"]).strip(): row for row in coord_rows}
if set(coord_by_id) != {agent["id"] for agent in agents}:
    raise RuntimeError("Kode agen workbook koordinat tidak identik dengan master agen kanonis")
for agent in agents:
    row = coord_by_id[agent["id"]]
    agent.update({
        "address": row["Alamat Lokasi / Alamat Kerja"] or agent["address"],
        "desaKelurahan": row["Kelurahan/Desa"], "kecamatan": row["Kecamatan"],
        "latitude": float(row["Latitude Kandidat"]), "longitude": float(row["Longitude Kandidat"]),
        "plusCode": row["Plus Code"], "googleMapsUrl": row["Link Pencarian Google Maps"],
        "coordinateStatus": "KANDIDAT_PERLU_VERIFIKASI_FAKTUAL", "gpsVerified": False,
        "pointType": "AGENT_LOCATION_CANDIDATE", "coordinateSource": row["Sumber Data Kandidat"],
        "coordinateNote": row["Catatan Verifikasi"],
        "locationVerification": {"status": "CANDIDATE_REQUIRES_FIELD_VERIFICATION", "verifiedAt": None, "verifiedBy": None},
        "dataVersion": VERSION,
    })

agent_id = {normalized(a["name"]): a["id"] for a in agents}
pangkalan = []
for row in geo_rows:
    precision = str(row["Level Presisi"])
    exact = precision == "A" and row["Status Match"] == "MATCH"
    partial = precision == "B"
    pid = f"PG-{int(row['No']):06d}"
    latitude=float(row["Latitude"]); longitude=float(row["Longitude"])
    coordinate_source=row["Sumber Koordinat Administratif"]
    correction_note=None
    # Referensi nonresmi workbook untuk Sipatuo jatuh di polygon Batulappa.
    # Gunakan representative point polygon Sipatuo dari overlay BIG lokal.
    if clean_village(row["Desa/Kelurahan"]).casefold()=="sipatuo" and row["Kecamatan (Normalisasi)"]=="Patampanua":
        latitude=-3.69821; longitude=119.7075861
        coordinate_source="BIG/Satu Data Pinrang — representative point polygon Desa Sipatuo (batas indikatif)"
        correction_note="Referensi workbook dikoreksi karena berada di luar polygon Patampanua; tetap merupakan fallback area, bukan GPS pangkalan."
    item = {
        "id": pid, "agentId": agent_id[normalized(row["Nama Agen/Penyalur"])],
        "agentName": row["Nama Agen/Penyalur"], "name": row["Nama Pangkalan/Subpenyalur"],
        "normalizedName": normalized(row["Nama Pangkalan/Subpenyalur"]),
        "kecamatan": row["Kecamatan (Normalisasi)"], "kecamatanSource": row["Kecamatan (Sumber)"],
        "desaKelurahan": clean_village(row["Desa/Kelurahan"]),
        "desaKelurahanSource": row["Desa/Kelurahan"], "address": row["Alamat ESDM"],
        "latitude": latitude, "longitude": longitude,
        "googleMapsUrl": row["Google Maps URL"], "googleMapsSearchUrl": row["Google Maps Search URL (Alamat ESDM)"],
        "plusCode": row["Plus Code"], "listingName": row["Nama Listing Maps"],
        "locationPrecisionLevel": precision, "locationMatchStatus": row["Status Match"],
        "locationMethod": row["Metode Lokasi"], "locationConfidence": row["Confidence"],
        "pointType": row["Jenis Titik Peta"], "gpsVerified": False,
        "locationVerification": {
            "status": "DIGITAL_LISTING_MATCH" if exact else ("DIGITAL_PARTIAL_MATCH" if partial else "ADMIN_AREA_FALLBACK"),
            "requiresAgentGps": row["Perlu Verifikasi Agen"] == "YA",
            "note": row["Catatan Verifikasi"], "checkedAt": "2026-08-31",
        },
        "verificationStatus": "PENDING_ADMIN_VERIFICATION",
        "status": "ACTIVE", "isDeleted": False,
        "sourceType": "DITJEN_MIGAS_Q1_2026_WITH_GEOSPATIAL_FALLBACK",
        "sourceDate": "2026-03-31", "sourceUrl": row["Sumber ESDM URL"],
        "sourceLine": int(row["Ref Line PDF"]), "geolocationSource": row["Sumber Geolokasi"],
        "administrativeCoordinateSource": coordinate_source,
        "coordinateCorrectionNote": correction_note,
        "dataVersion": VERSION,
        "sourceOriginal": {"kecamatan": row["Kecamatan (Sumber)"], "kelurahan": row["Desa/Kelurahan"], "namaSubPenyalur": row["Nama Pangkalan/Subpenyalur"], "alamatSubPenyalur": row["Alamat ESDM"], "namaPenyalur": row["Nama Agen/Penyalur"]},
    }
    pangkalan.append(item)

payload = (
    "// FALLBACK SNAPSHOT LPG — sumber utama runtime: Firestore lpg_agents/lpg_pangkalan\n"
    f"const LPG_DATA_VERSION = {json.dumps(VERSION)};\n"
    "const LPG_SEED_AGENTS = " + json.dumps(agents, ensure_ascii=False, indent=2) + ";\n\n"
    "const LPG_SEED_PANGKALAN = " + json.dumps(pangkalan, ensure_ascii=False, indent=2) + ";\n\n"
    "if (typeof window !== 'undefined') { window.LPG_DATA_VERSION=LPG_DATA_VERSION; window.LPG_SEED_AGENTS=LPG_SEED_AGENTS; window.LPG_SEED_PANGKALAN=LPG_SEED_PANGKALAN; }\n"
)
(ROOT / "js" / "lpg-data-seed.js").write_text(payload, encoding="utf-8")
(ROOT / "assets" / "data" / "lpg-agents.json").write_text(json.dumps(agents, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
(ROOT / "assets" / "data" / "lpg-pangkalan.json").write_text(json.dumps(pangkalan, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
print(f"Imported {len(agents)} agents and {len(pangkalan)} pangkalan; version={VERSION}")
