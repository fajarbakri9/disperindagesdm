"""Merge the verified market workbook into the website market master data."""
from pathlib import Path
from openpyxl import load_workbook
import json, re, sys

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"D:\# DOWNLOAD\Master_Data_Pasar_Pinrang_Google_Maps_2026-08-31.xlsx")
ALIASES = {
    "pasar-sentral-pinrang": "Pasar Sentral Pinrang", "pasar-pekkabata": "Pasar Pekkabata",
    "pasar-bungi": "Pasar Bungi", "pasar-langnga": "Pasar Rakyat Langnga",
    "pasar-teppo-benteng": "Pasar Teppo", "pasar-kariango": "Pasar Kariango",
    "pasar-pajalele": "Pasar Pajalele", "pasar-marawi": "Pasar Marawi",
    "pasar-kampoeng-djaja": "Pasar Kampoeng Djaja", "pasar-cempa": "Pasar Cempa",
    "pasar-leppangang": "Pasar Leppangang", "pasar-lanrisang-jampue": "Pasar Lanrisang",
    "pasar-malimpung": "Pasar Malimpung", "pasar-labolong": "Pasar Labolong",
    "pasar-tuppu": "Pasar Tuppu", "pasar-paleteang": "Pasar Paleteang",
}

def load_existing():
    text=(ROOT/'js/markets-data.js').read_text(encoding='utf-8')
    return json.loads(re.search(r'const PINRANG_ALL_MARKETS = (\[.*\]);', text, re.S).group(1))

def rows():
    sheet=load_workbook(SOURCE, read_only=True, data_only=True)['Master Pasar']
    headers=[c.value for c in sheet[5]]
    return {d['Nama Database']:d for values in sheet.iter_rows(min_row=6, values_only=True)
            if values[0] for d in [dict(zip(headers, values))]}

def schedule_days(raw):
    if not raw: return []
    prefix=re.split(r'\s+\d{1,2}:|\s+±\d{1,2}:', raw, maxsplit=1)[0]
    return [x.strip() for x in prefix.replace('Listing menunjukkan ', '').split('/') if x.strip()]

data=load_existing(); master=rows(); matched=set()
for item in data:
    name=ALIASES.get(item['id'])
    row=master.get(name) if name else None
    if not row:
        item['masterDataStatus']='SUPPLEMENTAL_PERLU_VERIFIKASI'
        item['masterDataNote']='Tidak tercantum pada workbook Google Maps 31 Agustus 2026.'
        continue
    matched.add(name)
    item.update({
        'nama': row['Nama Database'], 'namaGoogleMaps': row['Nama di Google Maps'],
        'namaAlternatif': row['Alias'] or item.get('namaAlternatif'),
        'jenisPasar': row['Jenis Pasar'], 'kategoriPengelolaan': row['Kategori Pengelolaan'],
        'unitPengelola': row['UPTD / Pengelola'], 'statusKewenangan': row['Status Kewenangan'],
        'kecamatan': row['Kecamatan'], 'desaKelurahan': row['Desa/Kelurahan'],
        'alamat': row['Alamat Google Maps'], 'plusCode': (row['Plus Code'] or '').strip() or None,
        'latitude': row['Latitude'], 'longitude': row['Longitude'],
        'statusKoordinat': ('Titik tengah sel Plus Code — perlu validasi GPS pintu masuk'
                            if row['Status Koordinat'] == 'Tinggi' else (row['Status Koordinat'] or 'Belum terverifikasi')),
        'metodeKoordinat': 'PLUS_CODE_CELL_CENTER' if row['Plus Code'] else 'BELUM_DITETAPKAN',
        'jadwalGoogleMaps': row['Hari/Jam Google Maps'],
        'hariPasar': schedule_days(row['Hari/Jam Google Maps']) or item.get('hariPasar', []),
        'googleMapsUrl': row['Link Profil/Foto Google Maps'], 'dasarReferensiMaster': row['Dasar/Referensi'],
        'tanggalVerifikasi': '31 Agustus 2026', 'tahunData': '2026',
        'masterDataStatus': 'TERVERIFIKASI_WORKBOOK_GOOGLE_MAPS',
        'masterDataSource': SOURCE.name,
    })

missing=set(master)-matched
if missing: raise SystemExit(f'Workbook rows were not matched: {sorted(missing)}')
payload=json.dumps(data, ensure_ascii=False, indent=2)
(ROOT/'assets/data/markets.json').write_text(payload+'\n', encoding='utf-8')
(ROOT/'js/markets-data.js').write_text('// MASTER DATA PASAR KABUPATEN PINRANG — GOOGLE MAPS 31 AGUSTUS 2026\nconst PINRANG_ALL_MARKETS = '+payload+';\n\nif (typeof window !== "undefined") window.PINRANG_ALL_MARKETS = PINRANG_ALL_MARKETS;\n', encoding='utf-8')
print(f'Merged {len(matched)} workbook rows; retained {len(data)-len(matched)} supplemental row(s).')
